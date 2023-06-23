import numpy as np
import openpyxl as xl


class FourPole:
    A, Y, Z = 'a', 'y', 'z'

    def __init__(self, *args):

        if len(args) == 3:
            # print(args)
            # print('параметры',*self.params, sep = "\n")
            self.form = args[0]
            self.freqs = args[1]
            self.params = args[2]

            # print("инициализация")
            # print(self.params)
            # print(self.freqs)
            # print('параметры',self.params, sep = "\n")

        # иначе исключение

    def calc_az(self):
        #if self.form != self.__y:
            for i in range(len(self.params)):
                m = self.params[i]
                x10 = m[1, 0]
                m[0, 1] = np.linalg.det(m)
                m[1, 0] = 1
                m /= x10

                self.params[i] = np.round(m, 4)

            self.form = self.Z if self.form == self.A else self.A

    def calc_yz(self):
        #if self.form != self.__a:
            for i in range(len(self.params)):
                m = self.params[i]
                det = np.linalg.det(m)
                m = np.array(([m[1, 1], -m[0, 1]], [-m[1, 0], m[0, 0]])) / det

                self.params[i] = np.round(m, 4)
            print("До преобразования", self.form)
            self.form = self.Y if self.form == self.Z else self.Z
            print("После преобразования", self.form)

    def calc_ay(self):
        #if self.form != self.__z:
            for i in range(len(self.params)):
                m = self.params[i]
                det = np.linalg.det(m)

                if self.form == self.A:
                    m = np.array(([m[1, 1], -det], [-1, m[0, 0]])) / m[0, 1]
                else:
                    m = np.array(([m[1, 1], 1], [det, m[0, 0]])) / (-m[1, 0])

                self.params[i] = np.round(m, 3)

            self.form = self.Y if self.form == self.A else self.A

    def calc_form(self, f):
        if self.form != f:
            if self.form != self.A and f != self.A:
                self.calc_yz()
            elif self.form != self.Y and f != self.Y:
                self.calc_az()
            elif self.form != self.Z and f != self.Z:
                self.calc_ay()


    def get_labels(self):
        if self.form == self.A:
            return 'A', 'B', 'C', 'D' #'A (1)', 'B (Ом)', 'C (См)', 'D (1)'
        elif self.form == self.Y:
            return 'Y11', 'Y12', 'Y21', 'Y22' #'Y11 (См)', 'Y12 (См)', 'Y21 (См)', 'Y22 (См)'
        elif self.form == self.Z:
            return 'Z11', 'Z12', 'Z21', 'Z22' #'Z11 (Ом)', 'Z12 (Ом)', 'Z21 (Ом)', 'Z22 (Ом)'

    def get_coefs(self):
        params = self.params
        return params.reshape(len(params), 4).T.round(4)

    def get_mods(self):
        return np.absolute(self.get_coefs()).round(4)

    def get_phases(self):
        return np.angle(self.get_coefs()).round(4)

    def get_counts(self, ):
        return [2*np.pi/a for a in reversed(self.freqs)]

    def get_freqs_aug(self):
        freqs = self.freqs
        sample_len = freqs[0] if len(freqs) == 1 else freqs[1]-freqs[0]
        freqs_start = list(range(0, freqs[0], sample_len))
        return freqs_start+freqs

    def get_time_scale(self, freqs_values):
        fs = 2*max(freqs_values)
        ts = round(1/fs, 4)
        print('Период дискретизации: ',ts)
        time_values = [val*ts for val in range(0, len(freqs_values))]
        return time_values

    def get_ifft(self):
        params = self.params
        coefs = self.get_coefs()
        freqs_aug = self.get_freqs_aug()
        ifft = np.repeat(params.reshape(len(params), 4)[0].reshape(-1, 1), len(freqs_aug), axis=1) # матрица, инициализированная первыми значениями коэффициентов
        ifft[:, len(freqs_aug)-len(self.freqs):] = coefs

        return np.fft.ifft(ifft).round(3)
        #return np.fft.ifft(self.get_coefs())#FourPole.get_spaced_2d_h4(self.get_coefs(), k))



def calc_serial(pole1, pole2):  # посл-посл Z = Z1 + Z2
    return FourPole('z', pole1.freqs, pole1.params + pole2.params)


def calc_parallel(pole1, pole2):  # посл-парал Y = Y1 + Y2
    return FourPole('y', pole1.freqs, pole1.params + pole2.params)


def calc_cascade(pole1, pole2):  # каскадное A = A1*A2 матр
    return FourPole('a', pole1.freqs, pole1.params @ pole2.params)


def calc_con(pole1, pole2):
    if pole1.form == pole2.form:
        func = None
        f = pole1.form
        if f == 'a':
            func = calc_cascade
        elif f == 'y':
            func = calc_parallel
        elif f == 'z':
            func = calc_serial
        return func(pole1, pole2)

    print("Формы четырехполюсников не совпадают")
    return None

# cascPole = getCascade(pole1, pole2)

def get_data_from_excel(file_path):
    wb = xl.load_workbook(file_path)
    sheet1 = wb.worksheets[0]
    sheet2 = wb.worksheets[1]

    data1 = {}
    for row in sheet1.iter_rows(values_only=True):
        key = row[0]
        values = tuple(map(complex, row[1:]))
        data1[key] = values

    data2 = {}
    for row in sheet2.iter_rows(values_only=True):
        key = row[0]
        values = tuple(map(complex, row[1:]))
        data2[key] = values

    wb.close()

    return data1, data2
